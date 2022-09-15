import openpyxl as xl
from openpyxl.chart import BarChart, Reference

workBook = xl.load_workbook("transactions.xlsx")
sheet = workBook["Sheet1"]
"""
two ways of making a cell object
"""
cell = sheet["a1"]
cell = sheet.cell(1, 1)
print(cell.value)
print(sheet.max_row)  # checking how many rows in this xl sheet

# iterating through the sheet
for rows in range(1, sheet.max_row + 1):
    print(rows)

# iterating and accessing values from 3rd column of this xl file:
for rows in range(2, sheet.max_row + 1):
    cell = sheet.cell(rows, 3)  # this method takes rows value and number of column
    print(cell.value)
    # multiplying with each row with 0.9 in column 3:
    corrected_price = cell.value * 0.9
    # creating cell or column 4:
    corrected_price_cell = sheet.cell(rows, 4)
    # assigning values to 4th column:
    corrected_price_cell.value = corrected_price
# creating a chart:
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')
workBook.save("transactions.xlsx")
